"""
Template Cleanup Script for AquaExport Pro 2.0
==============================================
This script cleans up the template.xlsx file by removing any existing data
while preserving the structure and formatting needed for water quality exports.

Usage:
    python cleanup_template.py

The script will:
1. Remove all data from data cells while keeping headers and structure
2. Clear year values from year cells
3. Ensure proper sheet structure exists
4. Preserve formatting and formulas
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.utils import get_column_letter
import os
import sys
from pathlib import Path

class TemplateCleaner:
    """Clean up Excel template for water quality data export."""
    
    # Excel layout constants (matching exporter.py)
    MONTH_NAMES = {
        1: "siječanj", 2: "veljača", 3: "ožujak", 4: "travanj",
        5: "svibanj", 6: "lipanj", 7: "srpanj", 8: "kolovoz",
        9: "rujan", 10: "listopad", 11: "studeni", 12: "prosinac"
    }
    
    BLOCK_ANCHORS = {
        'PK Barbat': 11,
        'VS Lopar': 59,
        'VS Perici': 107
    }
    
    # Year cells that should be cleared
    YEAR_CELLS = ['B9', 'B57', 'B105']
    
    def __init__(self, template_path="template.xlsx"):
        self.template_path = Path(template_path)
        self.wb = None
        
    def load_workbook(self):
        """Load the Excel workbook."""
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template file not found: {self.template_path}")
        
        print(f"Loading template: {self.template_path}")
        self.wb = openpyxl.load_workbook(self.template_path)
        print(f"✓ Loaded workbook with {len(self.wb.sheetnames)} sheets")
        
    def create_backup(self):
        """Create a backup of the original template."""
        backup_path = self.template_path.with_suffix('.backup.xlsx')
        if not backup_path.exists():
            import shutil
            shutil.copy2(self.template_path, backup_path)
            print(f"✓ Created backup: {backup_path}")
        else:
            print(f"⚠ Backup already exists: {backup_path}")
            
    def ensure_sheet_structure(self):
        """Ensure all required monthly sheets exist."""
        print("\nChecking sheet structure...")
        
        # Check for required monthly sheets
        required_sheets = [f"P-{month:02d}" for month in range(1, 13)]
        missing_sheets = []
        
        for sheet_name in required_sheets:
            if sheet_name not in self.wb.sheetnames:
                missing_sheets.append(sheet_name)
                # Create the sheet
                ws = self.wb.create_sheet(sheet_name)
                print(f"  ✓ Created missing sheet: {sheet_name}")
                self.setup_monthly_sheet(ws, sheet_name)
        
        if not missing_sheets:
            print("  ✓ All required sheets exist")
            
        # Remove any extra sheets that aren't monthly sheets
        sheets_to_remove = []
        for sheet_name in self.wb.sheetnames:
            if not sheet_name.startswith('P-') and sheet_name != 'Sheet':
                sheets_to_remove.append(sheet_name)
                
        for sheet_name in sheets_to_remove:
            del self.wb[sheet_name]
            print(f"  ✓ Removed extra sheet: {sheet_name}")
            
    def setup_monthly_sheet(self, ws, sheet_name):
        """Set up a new monthly sheet with proper structure."""
        month_num = int(sheet_name.split('-')[1])
        month_name = self.MONTH_NAMES[month_num]
        
        # Set up headers and structure
        ws['A1'] = f"Podaci kvalitete vode - {month_name.upper()}"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Set up location headers
        for i, (location, anchor) in enumerate(self.BLOCK_ANCHORS.items()):
            # Location name
            ws[f'B{anchor}'] = location
            ws[f'B{anchor}'].font = Font(bold=True)
            
            # Day column header
            ws[f'A{anchor + 1}'] = "Dan"
            ws[f'A{anchor + 1}'].font = Font(bold=True)
            
            # Add day numbers (1-31) for each location block
            for day in range(1, 32):
                row = anchor + 2 + (day - 1)
                ws[f'A{row}'] = day
                ws[f'A{row}'].font = Font(bold=True)
            
            # Parameter headers based on location
            if location == 'PK Barbat':
                # PK Barbat has all 5 parameters
                headers = [
                    ('C', 'Mutnoca', 'Max'), ('D', 'Mutnoca', 'Min'), ('E', 'Mutnoca', 'Avg'),
                    ('F', 'Klor', 'Max'), ('G', 'Klor', 'Min'), ('H', 'Klor', 'Avg'),
                    ('I', 'Temp', 'Max'), ('J', 'Temp', 'Min'), ('K', 'Temp', 'Avg'),
                    ('L', 'pH', 'Max'), ('M', 'pH', 'Min'), ('N', 'pH', 'Avg'),
                    ('O', 'Redox', 'Max'), ('P', 'Redox', 'Min'), ('Q', 'Redox', 'Avg')
                ]
            else:
                # VS Lopar and VS Perici only have 3 parameters
                headers = [
                    ('C', 'Klor', 'Max'), ('D', 'Klor', 'Min'), ('E', 'Klor', 'Avg'),
                    ('F', 'Temp', 'Max'), ('G', 'Temp', 'Min'), ('H', 'Temp', 'Avg'),
                    ('I', 'Redox', 'Max'), ('J', 'Redox', 'Min'), ('K', 'Redox', 'Avg')
                ]
            
            # Write headers
            for col, param, stat in headers:
                cell = ws[f'{col}{anchor + 1}']
                cell.value = f"{param}\n{stat}"
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
        # Set up year cells
        for cell_ref in self.YEAR_CELLS:
            ws[cell_ref] = ""  # Clear year value
            ws[cell_ref].font = Font(bold=True)
            
        print(f"  ✓ Set up structure for {sheet_name}")
        
    def clean_data_cells(self):
        """Remove data from all data cells while preserving structure."""
        print("\nCleaning data cells...")
        
        for sheet_name in self.wb.sheetnames:
            if not sheet_name.startswith('P-'):
                continue
                
            ws = self.wb[sheet_name]
            print(f"  Cleaning {sheet_name}...")
            
            # Clear year cells
            for cell_ref in self.YEAR_CELLS:
                if ws[cell_ref].value and str(ws[cell_ref].value).isdigit():
                    ws[cell_ref].value = ""
                    print(f"    ✓ Cleared year cell {cell_ref}")
            
            # Clear data cells for each location block
            for location, anchor in self.BLOCK_ANCHORS.items():
                # Determine column range based on location
                if location == 'PK Barbat':
                    # PK Barbat has all 5 parameters (C through Q)
                    data_cols = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']
                else:
                    # VS Lopar and VS Perici only have 3 parameters (C through K)
                    data_cols = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
                
                # Clear data rows (days 1-31)
                for day in range(1, 32):
                    row = anchor + 2 + (day - 1)
                    
                    # Keep day number - don't clear it
                    # if ws[f'A{row}'].value and str(ws[f'A{row}'].value).isdigit():
                    #     ws[f'A{row}'].value = ""
                        
                    # Clear location name (only on first day)
                    if day == 1 and ws[f'B{row}'].value == location:
                        ws[f'B{row}'].value = location  # Keep the location name
                    
                    # Clear data cells
                    for col in data_cols:
                        cell = ws[f'{col}{row}']
                        if cell.value is not None and cell.value != "":
                            # Check if it's a number (data) vs text (header)
                            try:
                                float(cell.value)
                                cell.value = ""  # Clear numeric data
                            except (ValueError, TypeError):
                                # Keep non-numeric values (headers, etc.)
                                pass
                                
            print(f"    ✓ Cleaned data cells for {sheet_name}")
            
    def clean_day_numbers(self):
        """Preserve day numbers in column A."""
        print("\nPreserving day numbers...")
        
        for sheet_name in self.wb.sheetnames:
            if not sheet_name.startswith('P-'):
                continue
                
            ws = self.wb[sheet_name]
            
            for location, anchor in self.BLOCK_ANCHORS.items():
                for day in range(1, 32):
                    row = anchor + 2 + (day - 1)
                    cell = ws[f'A{row}']
                    
                    # Keep day numbers - don't clear them
                    # if cell.value and str(cell.value).isdigit():
                    #     cell.value = ""
                        
            print(f"  ✓ Preserved day numbers in {sheet_name}")
            
    def preserve_formatting(self):
        """Ensure proper formatting is preserved."""
        print("\nPreserving formatting...")
        
        for sheet_name in self.wb.sheetnames:
            if not sheet_name.startswith('P-'):
                continue
                
            ws = self.wb[sheet_name]
            
            # Ensure proper column widths
            column_widths = {
                'A': 8,   # Day
                'B': 15,  # Location
                'C': 12,  # Data columns
                'D': 12,
                'E': 12,
                'F': 12,
                'G': 12,
                'H': 12,
                'I': 12,
                'J': 12,
                'K': 12,
                'L': 12,
                'M': 12,
                'N': 12,
                'O': 12,
                'P': 12,
                'Q': 12,
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
                
            print(f"  ✓ Preserved formatting in {sheet_name}")
            
    def save_workbook(self):
        """Save the cleaned workbook."""
        print(f"\nSaving cleaned template...")
        self.wb.save(self.template_path)
        print(f"✓ Saved cleaned template: {self.template_path}")
        
    def cleanup(self):
        """Perform complete template cleanup."""
        print("🧹 AquaExport Template Cleanup")
        print("=" * 40)
        
        try:
            # Create backup
            self.create_backup()
            
            # Load workbook
            self.load_workbook()
            
            # Ensure proper structure
            self.ensure_sheet_structure()
            
            # Clean data cells
            self.clean_data_cells()
            
            # Clear day numbers
            self.clean_day_numbers()
            
            # Preserve formatting
            self.preserve_formatting()
            
            # Save workbook
            self.save_workbook()
            
            print("\n✅ Template cleanup completed successfully!")
            print(f"   Original backed up as: {self.template_path.with_suffix('.backup.xlsx')}")
            print(f"   Cleaned template: {self.template_path}")
            
        except Exception as e:
            print(f"\n❌ Error during cleanup: {e}")
            return False
            
        return True

def main():
    """Main entry point."""
    if len(sys.argv) > 1:
        template_path = sys.argv[1]
    else:
        template_path = "template.xlsx"
        
    cleaner = TemplateCleaner(template_path)
    success = cleaner.cleanup()
    
    if success:
        print("\n🎉 Template is now ready for use!")
        print("   The template has been cleaned of all data while preserving structure.")
        print("   You can now use it with AquaExport Pro 2.0.")
    else:
        print("\n💥 Cleanup failed. Please check the error messages above.")
        sys.exit(1)

if __name__ == "__main__":
    main()