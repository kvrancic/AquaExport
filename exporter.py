"""
AquaExport Pro 2.0 - Dual Mode Water Data Exporter
==================================================
A modern, high-performance tool for exporting water quality and quantity data from PostgreSQL to Excel.

Features:
- Dual mode: Water Quality (Kvaliteta vode) and Water Quantities (Zahvaćene količine vode)
- Beautiful GUI with date pickers and mode selection
- Dynamic color themes (blue for quality, green for quantities)
- Blazing fast performance using bulk operations
- Idempotent file handling (creates or updates yearly workbooks)
- Comprehensive error handling and logging
- Single-file executable support

Author: Water Data Export System
Version: 2.1.0
"""

import os
import sys
import logging
import traceback
import asyncio
import shutil
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any, Literal
from dataclasses import dataclass
from collections import defaultdict
from enum import Enum
import tomli
import psycopg2
from psycopg2.extras import RealDictCursor
import openpyxl
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
import threading


class ExportMode(Enum):
    """Export mode enumeration."""
    WATER_QUALITY = "kvaliteta_vode"
    WATER_QUANTITIES = "zahvacene_kolicine_vode"


# Configure logging
def setup_logging(export_dir: Path) -> logging.Logger:
    """Set up rotating log file in export directory."""
    log_file = export_dir / "exporter.log"

    # Create formatter
    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )

    # File handler with rotation
    from logging.handlers import RotatingFileHandler
    file_handler = RotatingFileHandler(
        log_file, maxBytes=10 * 1024 * 1024, backupCount=5
    )
    file_handler.setFormatter(formatter)

    # Console handler
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)

    # Set up logger
    logger = logging.getLogger('AquaExportPro')
    logger.setLevel(logging.DEBUG)
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger


@dataclass
class TagMapping:
    """Maps database tag indices to Excel columns for each location."""
    location: str
    mutnoca: Optional[int] = None
    klor: Optional[int] = None
    temp: Optional[int] = None
    pH: Optional[int] = None
    redox: Optional[int] = None


@dataclass
class QuantityTagMapping:
    """Maps database tag indices for water quantity measurements."""
    location: str
    daily_volume_in: Optional[int] = None
    daily_volume_out: Optional[int] = None
    max_flow_in: Optional[int] = None
    max_flow_out: Optional[int] = None


@dataclass
class Config:
    """Application configuration."""
    db_host: str
    db_port: int
    db_name: str
    db_user: str
    db_password: str
    export_dir: Path
    template_dir: Path

    @classmethod
    def from_file(cls, config_path: str = "config.toml") -> 'Config':
        """Load configuration from TOML file."""
        try:
            with open(config_path, "rb") as f:
                data = tomli.load(f)

            # Handle backward compatibility
            template_path = data["export"].get("template_path")
            template_dir = data["export"].get("template_dir", "./templates")

            if template_path and not Path(template_dir).exists():
                # Migrate from old single template to new structure
                template_dir = Path("./templates")

            return cls(
                db_host=data["database"]["host"],
                db_port=data["database"]["port"],
                db_name=data["database"]["name"],
                db_user=data["database"]["user"],
                db_password=data["database"]["password"],
                export_dir=Path(data["export"]["directory"]),
                template_dir=Path(template_dir)
            )
        except Exception as e:
            # Default configuration if file not found
            return cls(
                db_host="localhost",
                db_port=5432,
                db_name="SCADA_arhiva_rab",
                db_user="postgres",
                db_password="fornax123",
                export_dir=Path("./exports"),
                template_dir=Path("./templates")
            )


class WaterDataExporter:
    """Main exporter class handling database queries and Excel writing for both modes."""

    # Excel layout constants
    MONTH_NAMES = {
        1: "siječanj", 2: "veljača", 3: "ožujak", 4: "travanj",
        5: "svibanj", 6: "lipanj", 7: "srpanj", 8: "kolovoz",
        9: "rujan", 10: "listopad", 11: "studeni", 12: "prosinac"
    }

    # Water Quality block anchors
    QUALITY_BLOCK_ANCHORS = {
        'PK Barbat': 11,
        'VS Lopar': 59,
        'VS Perici': 107
    }

    # Water Quantities block anchors (same for all sheets)
    QUANTITY_BLOCK_ANCHORS = {
        'Hrvatsko primorje južni ogranak': 11,
        'Perići': 60,
        'Gvačići I': 108,
        'Mlinica': 156,
        'Gvačići II': 204
    }

    def __init__(self, config: Config, logger: logging.Logger):
        self.config = config
        self.logger = logger
        self.conn = None

        # Load tag mappings from config for water quality
        self.quality_mappings = self._load_quality_mappings()

        # Define quantity mappings
        self.quantity_mappings = self._load_quantity_mappings()

    def _load_quality_mappings(self) -> Dict[str, TagMapping]:
        """Load water quality tag mappings from config file or use defaults."""
        default_mappings = {
            'PK Barbat': TagMapping(
                location='PK Barbat',
                mutnoca=3,
                klor=21,
                temp=134,
                pH=132,
                redox=133
            ),
            'VS Lopar': TagMapping(
                location='VS Lopar',
                klor=151,
                temp=155,
                redox=156
            ),
            'VS Perici': TagMapping(
                location='VS Perici',
                klor=72,
                temp=82,
                redox=81
            )
        }

        # Try to load from config file
        try:
            import tomli
            with open("config.toml", "rb") as f:
                config_data = tomli.load(f)

            if "tag_mappings" in config_data:
                self.logger.info("Loading quality tag mappings from config.toml")
                mappings = {}

                for location, tags in config_data["tag_mappings"].items():
                    location_name = {
                        'pk_barbat': 'PK Barbat',
                        'vs_lopar': 'VS Lopar',
                        'vs_perici': 'VS Perici'
                    }.get(location.lower(), location)

                    mappings[location_name] = TagMapping(
                        location=location_name,
                        mutnoca=tags.get('mutnoca'),
                        klor=tags.get('klor'),
                        temp=tags.get('temp'),
                        pH=tags.get('pH'),
                        redox=tags.get('redox')
                    )

                return mappings
        except Exception as e:
            self.logger.debug(f"Using default quality tag mappings: {e}")

        return default_mappings

    def _load_quantity_mappings(self) -> Dict[str, QuantityTagMapping]:
        """Load water quantity tag mappings."""
        return {
            'Hrvatsko primorje južni ogranak': QuantityTagMapping(
                location='Hrvatsko primorje južni ogranak',
                daily_volume_in=14,  # pk_barb\protok_dnevni_ul
                daily_volume_out=13,  # pk_barb\protok_dnevni_iz
                max_flow_in=18,  # pk_barb\protok_ulaz
                max_flow_out=16  # pk_barb\protok_izlaz
            ),
            'Perići': QuantityTagMapping(
                location='Perići',
                daily_volume_in=67,  # vs_perici\cs_protok_dnevni
                max_flow_in=68  # vs_perici\cs_protok_izlaz
            ),
            'Gvačići I': QuantityTagMapping(
                location='Gvačići I',
                daily_volume_in=103,  # bus_gvacici1\cs_protok_dnevni
                max_flow_in=0  # bus_gvacici1\cs_protok_izlaz
            ),
            'Mlinica': QuantityTagMapping(
                location='Mlinica',
                daily_volume_in=51,  # vs_fruga\MP_Mlinica_dnevni
                max_flow_in=52  # vs_fruga\MP_Mlinica_protok
            )
            # Gvačići II is skipped for now
        }

    def connect_db(self) -> None:
        """Establish database connection."""
        try:
            self.conn = psycopg2.connect(
                host=self.config.db_host,
                port=self.config.db_port,
                database=self.config.db_name,
                user=self.config.db_user,
                password=self.config.db_password,
                cursor_factory=RealDictCursor
            )
            self.logger.info("Database connection established")
        except Exception as e:
            self.logger.error(f"Database connection failed: {e}")
            raise

    def disconnect_db(self) -> None:
        """Close database connection."""
        if self.conn:
            self.conn.close()
            self.logger.info("Database connection closed")

    def fetch_quality_data(self, start_date: date, end_date: date) -> Dict[str, Dict]:
        """
        Fetch water quality data from database for date range.
        Returns nested dict: {location -> {date -> {param -> (min, max, avg)}}}
        """
        if not self.conn:
            self.connect_db()

        results = defaultdict(lambda: defaultdict(dict))

        try:
            with self.conn.cursor() as cur:
                query = """
                    SELECT 
                        date_trunc('day', dateandtime) AS day,
                        COALESCE(MIN(CASE WHEN val > 0 THEN val END), 0) AS min_val,
                        COALESCE(MAX(val), 0) AS max_val,
                        COALESCE(AVG(CASE WHEN val > 0 THEN val END), 0) AS avg_val
                    FROM floattable
                    WHERE dateandtime >= %s 
                        AND dateandtime < %s + INTERVAL '1 day'
                        AND tagindex = %s
                    GROUP BY day
                    ORDER BY day
                """

                for location, mapping in self.quality_mappings.items():
                    self.logger.info(f"Fetching quality data for {location}")

                    params = {
                        'mutnoca': mapping.mutnoca,
                        'klor': mapping.klor,
                        'temp': mapping.temp,
                        'pH': mapping.pH,
                        'redox': mapping.redox
                    }

                    for param_name, tag_index in params.items():
                        if tag_index is None:
                            continue

                        cur.execute(query, (start_date, end_date, tag_index))
                        rows = cur.fetchall()

                        for row in rows:
                            day = row['day'].date()
                            results[location][day][param_name] = (
                                round(row['min_val'], 2) if row['min_val'] else None,
                                round(row['max_val'], 2) if row['max_val'] else None,
                                round(row['avg_val'], 2) if row['avg_val'] else None
                            )

                        self.logger.debug(f"  {param_name}: {len(rows)} days of data")

        except Exception as e:
            self.logger.error(f"Error fetching quality data: {e}")
            raise

        return dict(results)

    def fetch_quantity_data(self, start_date: date, end_date: date) -> Dict[str, Dict]:
        """
        Fetch water quantity data from database for date range.
        Returns nested dict: {location -> {date -> {metric -> value}}}
        """
        if not self.conn:
            self.connect_db()

        results = defaultdict(lambda: defaultdict(dict))

        try:
            with self.conn.cursor() as cur:
                # Query for daily volumes (cumulative counters that reset at night)
                # Look for max value between 22:00 today and 03:00 tomorrow
                volume_query = """
                    WITH windowed AS (
                        SELECT
                            CASE
                                WHEN date_part('hour', dateandtime) >= 22
                                THEN date_trunc('day', dateandtime)
                                ELSE date_trunc('day', dateandtime) - INTERVAL '1 day'
                            END AS shift_date,
                            val
                        FROM floattable
                        WHERE tagindex = %s
                          AND dateandtime >= %s::date + INTERVAL '22 hours'
                          AND dateandtime < (%s::date + INTERVAL '1 day') + INTERVAL '3 hours'
                          AND (
                                date_part('hour', dateandtime) >= 22
                             OR date_part('hour', dateandtime) < 3
                          )
                    )
                    SELECT
                        shift_date AS day,
                        COALESCE(MAX(val), 0) AS max_val
                    FROM windowed
                    WHERE shift_date >= %s AND shift_date <= %s
                    GROUP BY shift_date
                    ORDER BY shift_date
                """

                # Query for max flow (instantaneous readings during the day)
                flow_query = """
                    SELECT 
                        date_trunc('day', dateandtime) AS day,
                        COALESCE(MAX(val), 0) AS max_val
                    FROM floattable
                    WHERE tagindex = %s
                        AND dateandtime >= %s
                        AND dateandtime < %s + INTERVAL '1 day'
                    GROUP BY day
                    ORDER BY day
                """

                for location, mapping in self.quantity_mappings.items():
                    self.logger.info(f"Fetching quantity data for {location}")

                    # Fetch daily volume (input)
                    if mapping.daily_volume_in:
                        cur.execute(volume_query, (
                            mapping.daily_volume_in,
                            start_date - timedelta(days=1),  # Start day before for 22:00
                            end_date,
                            start_date,
                            end_date
                        ))
                        for row in cur.fetchall():
                            day = row['day'].date()
                            results[location][day]['volume_in'] = round(row['max_val'], 0)

                    # Fetch daily volume (output) - only for main location
                    if mapping.daily_volume_out:
                        cur.execute(volume_query, (
                            mapping.daily_volume_out,
                            start_date - timedelta(days=1),
                            end_date,
                            start_date,
                            end_date
                        ))
                        for row in cur.fetchall():
                            day = row['day'].date()
                            results[location][day]['volume_out'] = round(row['max_val'], 0)

                    # Fetch max flow (input)
                    if mapping.max_flow_in:
                        cur.execute(flow_query, (
                            mapping.max_flow_in,
                            start_date,
                            end_date
                        ))
                        for row in cur.fetchall():
                            day = row['day'].date()
                            results[location][day]['max_flow_in'] = round(row['max_val'], 2)

                    # Fetch max flow (output) - only for main location
                    if mapping.max_flow_out:
                        cur.execute(flow_query, (
                            mapping.max_flow_out,
                            start_date,
                            end_date
                        ))
                        for row in cur.fetchall():
                            day = row['day'].date()
                            results[location][day]['max_flow_out'] = round(row['max_val'], 2)

                    self.logger.debug(f"  Fetched {len(results[location])} days of data")

        except Exception as e:
            self.logger.error(f"Error fetching quantity data: {e}")
            raise

        return dict(results)

    def get_or_create_workbook(self, year: int, mode: ExportMode) -> Tuple[openpyxl.Workbook, Path]:
        """Get existing workbook or create new one from template."""
        # Ensure export directory exists
        mode_dir = self.config.export_dir / mode.value
        mode_dir.mkdir(parents=True, exist_ok=True)

        # Template and workbook paths
        if mode == ExportMode.WATER_QUALITY:
            template_name = "kvaliteta_vode_template.xlsx"
            wb_name = f"kvaliteta_vode_{year}.xlsx"
            sheet_prefix = "P-"
        else:
            template_name = "zahvacene_kolicine_vode_template.xlsx"
            wb_name = f"zahvacene_kolicine_{year}.xlsx"
            sheet_prefix = "P2-"

        template_path = self.config.template_dir / template_name
        wb_path = mode_dir / wb_name

        if wb_path.exists():
            self.logger.info(f"Opening existing workbook: {wb_path}")
            try:
                wb = openpyxl.load_workbook(wb_path)
            except PermissionError:
                raise PermissionError(
                    f"Datoteka {wb_path.name} je trenutno otvorena u Excel-u!\n\n"
                    f"Molimo zatvorite datoteku u Excel-u i pokušajte ponovno."
                )
        else:
            self.logger.info(f"Creating new workbook from template for year {year}")
            if not template_path.exists():
                raise FileNotFoundError(f"Template not found: {template_path}")

            shutil.copy2(template_path, wb_path)
            wb = openpyxl.load_workbook(wb_path)

            # Pre-fill year in all sheets
            if mode == ExportMode.WATER_QUALITY:
                for month in range(1, 13):
                    sheet_name = f"{sheet_prefix}{month:02d}"
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        for cell in ['B9', 'B57', 'B105']:
                            ws[cell] = year

        return wb, wb_path

    def write_quality_to_excel(self, data: Dict[str, Dict], start_date: date, end_date: date,
                               progress_callback=None) -> None:
        """Write water quality data to Excel workbook."""
        # Group data by year
        years_data = defaultdict(lambda: defaultdict(dict))

        current = start_date
        while current <= end_date:
            for location, location_data in data.items():
                if current in location_data:
                    years_data[current.year][location][current] = location_data[current]
            current += timedelta(days=1)

        total_records = sum(
            len(loc_data) for year_data in years_data.values()
            for loc_data in year_data.values()
        )
        processed_records = 0

        for year, year_data in years_data.items():
            wb, wb_path = self.get_or_create_workbook(year, ExportMode.WATER_QUALITY)

            try:
                for location, location_data in year_data.items():
                    block_anchor = self.QUALITY_BLOCK_ANCHORS.get(location)
                    if not block_anchor:
                        self.logger.warning(f"Unknown location: {location}")
                        continue

                    for day_date, params in location_data.items():
                        month = day_date.month
                        day = day_date.day

                        sheet_name = f"P-{month:02d}"
                        if sheet_name not in wb.sheetnames:
                            self.logger.warning(f"Sheet {sheet_name} not found")
                            continue

                        ws = wb[sheet_name]
                        row = block_anchor + 2 + (day - 1)

                        ws[f'A{row}'] = day
                        if day == 1:
                            ws[f'B{row}'] = location

                        # Different column layouts per location
                        if location == 'PK Barbat':
                            col_mapping = {
                                'mutnoca': {'max': 'C', 'min': 'D', 'avg': 'E'},
                                'klor': {'max': 'F', 'min': 'G', 'avg': 'H'},
                                'temp': {'max': 'I', 'min': 'J', 'avg': 'K'},
                                'pH': {'max': 'L', 'min': 'M', 'avg': 'N'},
                                'redox': {'max': 'O', 'min': 'P', 'avg': 'Q'}
                            }
                        else:
                            col_mapping = {
                                'klor': {'max': 'C', 'min': 'D', 'avg': 'E'},
                                'temp': {'max': 'F', 'min': 'G', 'avg': 'H'},
                                'redox': {'max': 'I', 'min': 'J', 'avg': 'K'}
                            }

                        for param, values in params.items():
                            if param in col_mapping and values:
                                min_val, max_val, avg_val = values
                                if max_val is not None:
                                    ws[f"{col_mapping[param]['max']}{row}"] = max_val
                                if min_val is not None:
                                    ws[f"{col_mapping[param]['min']}{row}"] = min_val
                                if avg_val is not None:
                                    ws[f"{col_mapping[param]['avg']}{row}"] = avg_val

                        processed_records += 1
                        if progress_callback:
                            progress_callback(processed_records, total_records)

                wb.save(wb_path)
                self.logger.info(f"Saved workbook: {wb_path}")

            except Exception as e:
                self.logger.error(f"Error writing quality data to Excel: {e}")
                raise
            finally:
                wb.close()

    def write_quantity_to_excel(self, data: Dict[str, Dict], start_date: date, end_date: date,
                                progress_callback=None) -> None:
        """Write water quantity data to Excel workbook."""
        # Group data by year
        years_data = defaultdict(lambda: defaultdict(dict))

        current = start_date
        while current <= end_date:
            for location, location_data in data.items():
                if current in location_data:
                    years_data[current.year][location][current] = location_data[current]
            current += timedelta(days=1)

        total_records = sum(
            len(loc_data) for year_data in years_data.values()
            for loc_data in year_data.values()
        )
        processed_records = 0

        for year, year_data in years_data.items():
            wb, wb_path = self.get_or_create_workbook(year, ExportMode.WATER_QUANTITIES)

            try:
                for location, location_data in year_data.items():
                    block_anchor = self.QUANTITY_BLOCK_ANCHORS.get(location)
                    if not block_anchor:
                        self.logger.warning(f"Unknown location: {location}")
                        continue

                    for day_date, metrics in location_data.items():
                        month = day_date.month
                        day = day_date.day

                        sheet_name = f"P2-{month:02d}"
                        if sheet_name not in wb.sheetnames:
                            self.logger.warning(f"Sheet {sheet_name} not found")
                            continue

                        ws = wb[sheet_name]
                        row = block_anchor + 2 + (day - 1)

                        # Write day number
                        ws[f'A{row}'] = day

                        # Write location name on first day
                        if day == 1:
                            ws[f'B{row}'] = location

                        # Write metrics based on location
                        if location == 'Hrvatsko primorje južni ogranak':
                            # This location has both input and output
                            if 'volume_in' in metrics:
                                ws[f'C{row}'] = metrics['volume_in']
                            if 'max_flow_in' in metrics:
                                ws[f'E{row}'] = metrics['max_flow_in']
                            if 'volume_out' in metrics:
                                ws[f'F{row}'] = metrics['volume_out']
                            # Column D has formula, don't overwrite
                        else:
                            # Other locations only have input
                            if 'volume_in' in metrics:
                                ws[f'C{row}'] = metrics['volume_in']
                            if 'max_flow_in' in metrics:
                                ws[f'E{row}'] = metrics['max_flow_in']
                            # Column D has formula, don't overwrite

                        processed_records += 1
                        if progress_callback:
                            progress_callback(processed_records, total_records)

                wb.save(wb_path)
                self.logger.info(f"Saved workbook: {wb_path}")

            except Exception as e:
                self.logger.error(f"Error writing quantity data to Excel: {e}")
                raise
            finally:
                wb.close()


class ModernDualModeGUI:
    """Modern GUI for the Water Data Exporter with dual mode support."""

    def __init__(self, config: Config, logger: logging.Logger):
        self.config = config
        self.logger = logger
        self.exporter = WaterDataExporter(config, logger)

        self.root = tk.Tk()
        self.root.title("AquaExport Pro 2.1 - Izvoz Podataka o Vodi")
        self.root.geometry("650x800")
        self.root.resizable(False, False)

        # Set icon (if available)
        try:
            self.root.iconbitmap("icon.ico")
        except:
            pass

        # Apply modern theme
        self.style = ttk.Style()
        self.style.theme_use('clam')

        # Color schemes
        self.bg_color = "#f0f4f8"
        self.quality_color = "#2196F3"  # Blue
        self.quantity_color = "#4CAF50"  # Green
        self.error_color = "#f44336"

        # Current mode
        self.current_mode = tk.StringVar(value=ExportMode.WATER_QUALITY.value)

        self.root.configure(bg=self.bg_color)

        self.setup_ui()

    def setup_ui(self):
        """Create the user interface."""
        # Main container
        main_frame = tk.Frame(self.root, bg=self.bg_color)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        # Title
        self.add_header(main_frame)

        # Mode selection
        self.add_mode_selection(main_frame)

        # Date selection
        self.add_date_selection(main_frame)

        # Export button
        self.add_export_button(main_frame)

        # Progress section
        self.add_progress_section(main_frame)

        # Status/log section
        self.add_status_section(main_frame)

        # Footer
        self.add_footer(main_frame)

        # Help button
        self.add_help_button(main_frame)

        # Apply initial color scheme
        self.update_color_scheme()

    def add_header(self, parent):
        """Add title header."""
        header_frame = tk.Frame(parent, bg=self.bg_color)
        header_frame.pack(fill=tk.X, pady=(0, 20))

        title_label = tk.Label(
            header_frame,
            text="AquaExport Pro 2.1",
            font=("Segoe UI", 24, "bold"),
            fg=self.quality_color,
            bg=self.bg_color
        )
        title_label.pack(anchor=tk.W)
        self.title_label = title_label  # Store reference for color updates

        subtitle_label = tk.Label(
            header_frame,
            text="Izvoz podataka o vodi",
            font=("Segoe UI", 12),
            fg="#666",
            bg=self.bg_color
        )
        subtitle_label.pack(anchor=tk.W)

    def add_mode_selection(self, parent):
        """Add mode selection radio buttons."""
        mode_frame = tk.LabelFrame(
            parent,
            text="Odabir načina rada",
            font=("Segoe UI", 12, "bold"),
            bg=self.bg_color,
            fg="#333",
            relief=tk.FLAT,
            borderwidth=2
        )
        mode_frame.pack(fill=tk.X, pady=(0, 20))

        inner_frame = tk.Frame(mode_frame, bg=self.bg_color)
        inner_frame.pack(padx=20, pady=15)

        # Water Quality radio button
        quality_radio = tk.Radiobutton(
            inner_frame,
            text="KVALITETA VODE",
            variable=self.current_mode,
            value=ExportMode.WATER_QUALITY.value,
            font=("Segoe UI", 11, "bold"),
            bg=self.bg_color,
            fg=self.quality_color,
            activebackground=self.bg_color,
            selectcolor=self.bg_color,
            command=self.update_color_scheme
        )
        quality_radio.grid(row=0, column=0, padx=(0, 40), sticky=tk.W)

        # Water Quantities radio button
        quantity_radio = tk.Radiobutton(
            inner_frame,
            text="ZAHVAĆENE KOLIČINE VODE",
            variable=self.current_mode,
            value=ExportMode.WATER_QUANTITIES.value,
            font=("Segoe UI", 11, "bold"),
            bg=self.bg_color,
            fg=self.quantity_color,
            activebackground=self.bg_color,
            selectcolor=self.bg_color,
            command=self.update_color_scheme
        )
        quantity_radio.grid(row=0, column=1, sticky=tk.W)

        # Store references
        self.quality_radio = quality_radio
        self.quantity_radio = quantity_radio

    def add_date_selection(self, parent):
        """Add date selection widgets."""
        self.date_frame = tk.LabelFrame(
            parent,
            text="Odabir Datuma",
            font=("Segoe UI", 12, "bold"),
            bg=self.bg_color,
            fg=self.quality_color,
            relief=tk.FLAT,
            borderwidth=2
        )
        self.date_frame.pack(fill=tk.X, pady=(0, 20))

        inner_frame = tk.Frame(self.date_frame, bg=self.bg_color)
        inner_frame.pack(padx=20, pady=20)

        # Start date
        start_label = tk.Label(
            inner_frame,
            text="Početni datum:",
            font=("Segoe UI", 11),
            bg=self.bg_color
        )
        start_label.grid(row=0, column=0, sticky=tk.W, padx=(0, 10), pady=(0, 10))

        self.start_date = DateEntry(
            inner_frame,
            width=12,
            background=self.quality_color,
            foreground='white',
            borderwidth=2,
            font=("Segoe UI", 10),
            date_pattern='dd.mm.yyyy'
        )
        self.start_date.grid(row=0, column=1, padx=(0, 30), pady=(0, 10))

        # End date
        end_label = tk.Label(
            inner_frame,
            text="Završni datum:",
            font=("Segoe UI", 11),
            bg=self.bg_color
        )
        end_label.grid(row=0, column=2, sticky=tk.W, padx=(0, 10), pady=(0, 10))

        self.end_date = DateEntry(
            inner_frame,
            width=12,
            background=self.quality_color,
            foreground='white',
            borderwidth=2,
            font=("Segoe UI", 10),
            date_pattern='dd.mm.yyyy'
        )
        self.end_date.grid(row=0, column=3, pady=(0, 10))

        # Quick select buttons
        quick_frame = tk.Frame(inner_frame, bg=self.bg_color)
        quick_frame.grid(row=1, column=0, columnspan=4, pady=(10, 0))

        tk.Label(
            quick_frame,
            text="Brzi odabir:",
            font=("Segoe UI", 9),
            bg=self.bg_color,
            fg="#666"
        ).pack(side=tk.LEFT, padx=(0, 10))

        self.quick_buttons = []
        for text, days in [("Danas", 0), ("7 dana", 7), ("30 dana", 30), ("Godina", 365)]:
            btn = tk.Button(
                quick_frame,
                text=text,
                font=("Segoe UI", 9),
                bg="white",
                relief=tk.FLAT,
                borderwidth=1,
                padx=15,
                command=lambda d=days: self.set_date_range(d)
            )
            btn.pack(side=tk.LEFT, padx=2)
            self.quick_buttons.append(btn)

    def add_export_button(self, parent):
        """Add the main export button."""
        self.export_button = tk.Button(
            parent,
            text="IZVEZI PODATKE",
            font=("Segoe UI", 14, "bold"),
            bg=self.quality_color,
            fg="white",
            relief=tk.FLAT,
            height=2,
            command=self.start_export
        )
        self.export_button.pack(fill=tk.X, pady=(0, 20))

    def add_progress_section(self, parent):
        """Add progress bar and status."""
        progress_frame = tk.Frame(parent, bg=self.bg_color)
        progress_frame.pack(fill=tk.X, pady=(0, 10))

        self.progress_label = tk.Label(
            progress_frame,
            text="Spremno za izvoz",
            font=("Segoe UI", 10),
            bg=self.bg_color,
            fg="#666"
        )
        self.progress_label.pack(anchor=tk.W)

        self.progress_bar = ttk.Progressbar(
            progress_frame,
            mode='determinate',
            style="Custom.Horizontal.TProgressbar"
        )
        self.progress_bar.pack(fill=tk.X, pady=(5, 0))

    def add_status_section(self, parent):
        """Add status/log display."""
        status_frame = tk.LabelFrame(
            parent,
            text="Status",
            font=("Segoe UI", 10),
            bg=self.bg_color,
            relief=tk.FLAT,
            borderwidth=1
        )
        status_frame.pack(fill=tk.BOTH, expand=True)

        self.status_text = tk.Text(
            status_frame,
            height=8,
            font=("Consolas", 9),
            bg="white",
            relief=tk.FLAT,
            borderwidth=1
        )
        self.status_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Scrollbar
        scrollbar = ttk.Scrollbar(self.status_text)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.status_text.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.status_text.yview)

    def add_footer(self, parent):
        """Add company footer."""
        footer_frame = tk.Frame(parent, bg=self.bg_color)
        footer_frame.pack(fill=tk.X, pady=(10, 0))

        footer_label = tk.Label(
            footer_frame,
            text="Created by FORNAX d.o.o • info@fornax-automatika.hr",
            font=("Segoe UI", 8),
            fg="#999",
            bg=self.bg_color
        )
        footer_label.pack(anchor=tk.CENTER)

    def add_help_button(self, parent):
        """Add help button."""
        self.help_button = tk.Button(
            parent,
            text="?",
            font=("Segoe UI", 12, "bold"),
            bg=self.quality_color,
            fg="white",
            relief=tk.FLAT,
            width=3,
            height=1,
            command=self.show_help
        )
        self.help_button.place(relx=1.0, rely=0, anchor=tk.NE)

    def update_color_scheme(self):
        """Update UI colors based on selected mode."""
        mode = ExportMode(self.current_mode.get())
        theme_color = self.quality_color if mode == ExportMode.WATER_QUALITY else self.quantity_color

        # Update widget colors
        self.title_label.config(fg=theme_color)
        self.date_frame.config(fg=theme_color)
        self.start_date.config(background=theme_color)
        self.end_date.config(background=theme_color)
        self.export_button.config(bg=theme_color)
        self.help_button.config(bg=theme_color)

        # Update progress bar style
        self.style.configure(
            "Custom.Horizontal.TProgressbar",
            background=theme_color,
            troughcolor="#e0e0e0",
            borderwidth=0,
            lightcolor=theme_color,
            darkcolor=theme_color
        )

    def set_date_range(self, days):
        """Set date range for quick selection."""
        end = datetime.now().date()
        if days == 0:
            start = end
        elif days == 365:
            start = date(end.year, 1, 1)
        else:
            start = end - timedelta(days=days)

        self.start_date.set_date(start)
        self.end_date.set_date(end)

    def log_status(self, message, level="INFO"):
        """Add message to status display."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.status_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.status_text.see(tk.END)
        self.root.update()

    def update_progress(self, current, total):
        """Update progress bar."""
        if total > 0:
            progress = (current / total) * 100
            self.progress_bar['value'] = progress
            self.progress_label.config(text=f"Obrađeno: {current}/{total} zapisa")
            self.root.update()

    def start_export(self):
        """Start the export process in a separate thread."""
        self.export_button.config(state=tk.DISABLED, text="IZVOZ U TIJEKU...")
        self.progress_bar['value'] = 0
        self.status_text.delete(1.0, tk.END)

        # Run export in thread to keep GUI responsive
        thread = threading.Thread(target=self.run_export)
        thread.daemon = True
        thread.start()

    def run_export(self):
        """Run the actual export process."""
        try:
            start = self.start_date.get_date()
            end = self.end_date.get_date()
            mode = ExportMode(self.current_mode.get())

            if start > end:
                raise ValueError("Početni datum mora biti prije završnog datuma!")

            mode_text = "kvalitete vode" if mode == ExportMode.WATER_QUALITY else "zahvaćenih količina vode"
            self.log_status(f"Početak izvoza {mode_text}: {start} - {end}")

            # Connect to database
            self.log_status("Povezivanje s bazom podataka...")
            self.exporter.connect_db()

            # Fetch data based on mode
            self.log_status("Dohvaćanje podataka...")
            if mode == ExportMode.WATER_QUALITY:
                data = self.exporter.fetch_quality_data(start, end)
            else:
                data = self.exporter.fetch_quantity_data(start, end)

            if not data:
                self.log_status("Nema podataka za odabrani period!", "WARNING")
            else:
                total_records = sum(len(loc_data) for loc_data in data.values())
                self.log_status(f"Pronađeno {total_records} zapisa")

                # Write to Excel
                self.log_status("Pisanje u Excel...")
                if mode == ExportMode.WATER_QUALITY:
                    self.exporter.write_quality_to_excel(
                        data, start, end,
                        progress_callback=self.update_progress
                    )
                else:
                    self.exporter.write_quantity_to_excel(
                        data, start, end,
                        progress_callback=self.update_progress
                    )

                self.log_status("Izvoz završen uspješno!", "SUCCESS")

                # Show success message
                self.root.after(
                    0,
                    lambda: self.show_success_and_open_folder(mode)
                )

        except PermissionError as e:
            self.logger.error(f"Permission error: {e}")
            self.log_status(f"GREŠKA: Datoteka je otvorena u Excel-u!", "ERROR")

            self.root.after(
                0,
                lambda: messagebox.showerror(
                    "Datoteka je otvorena",
                    str(e)
                )
            )
        except Exception as e:
            self.logger.error(f"Export failed: {e}\n{traceback.format_exc()}")
            self.log_status(f"GREŠKA: {str(e)}", "ERROR")

            self.root.after(
                0,
                lambda: messagebox.showerror(
                    "Greška",
                    f"Izvoz nije uspio!\n\n{str(e)}\n\nDetalji su zapisani u log datoteku."
                )
            )

        finally:
            self.exporter.disconnect_db()
            self.root.after(0, self.reset_ui)

    def show_success_and_open_folder(self, mode: ExportMode):
        """Show success message and open the export folder."""
        mode_dir = self.config.export_dir / mode.value

        result = messagebox.showinfo(
            "Uspjeh",
            f"Podaci su uspješno izvezeni!\n\nLokacija: {mode_dir}\n\nKliknite OK da otvorite mapu."
        )

        if result == 'ok':
            try:
                os.startfile(str(mode_dir))
            except Exception as e:
                self.logger.warning(f"Could not open folder: {e}")
                messagebox.showinfo(
                    "Lokacija datoteka",
                    f"Datoteke su spremljene u:\n{mode_dir}"
                )

    def reset_ui(self):
        """Reset UI after export."""
        self.export_button.config(state=tk.NORMAL, text="IZVEZI PODATKE")
        self.progress_label.config(text="Spremno za izvoz")

    def show_help(self):
        """Show help dialog."""
        mode = ExportMode(self.current_mode.get())
        mode_specific = ""

        if mode == ExportMode.WATER_QUALITY:
            mode_specific = """
KVALITETA VODE:
   • Izvoz parametara kvalitete vode
   • Lokacije: PK Barbat, VS Lopar, VS Perici
   • Parametri: Mutnoća, Klor, Temperatura, pH, Redox
   • Datoteke: kvaliteta_vode_YYYY.xlsx"""
        else:
            mode_specific = """
ZAHVAĆENE KOLIČINE VODE:
   • Izvoz dnevnih količina i protoka
   • Lokacije: Hr. primorje, Perići, Gvačići I, Mlinica
   • Parametri: Dnevna količina (m³), Maks. protok (l/s)
   • Datoteke: zahvacene_kolicine_YYYY.xlsx"""

        help_text = f"""
AquaExport Pro 2.1 - Upute za korištenje

1. ODABIR NAČINA RADA:
   • Odaberite između dva načina izvoza
   • Boja sučelja se mijenja prema odabiru
{mode_specific}

2. ODABIR DATUMA:
   • Odaberite početni i završni datum
   • Koristite brze tipke za česte periode

3. IZVOZ PODATAKA:
   • Kliknite "IZVEZI PODATKE"
   • Pratite napredak u statusnoj traci

4. NAPOMENE:
   • Zatvorite Excel datoteke prije izvoza
   • Za probleme provjerite log datoteku
   • Kontakt: neven.vrancic@fornax-automatika.hr
        """

        messagebox.showinfo("Pomoć", help_text.strip())

    def run(self):
        """Start the GUI application."""
        self.root.mainloop()


def migrate_file_structure(config: Config, logger: logging.Logger):
    """Migrate from old file structure to new organized structure."""
    # Create new directory structure
    config.template_dir.mkdir(parents=True, exist_ok=True)
    (config.export_dir / ExportMode.WATER_QUALITY.value).mkdir(parents=True, exist_ok=True)
    (config.export_dir / ExportMode.WATER_QUANTITIES.value).mkdir(parents=True, exist_ok=True)

    # Migrate old template if exists
    old_template = Path("template.xlsx")
    new_quality_template = config.template_dir / "kvaliteta_vode_template.xlsx"

    if old_template.exists() and not new_quality_template.exists():
        logger.info(f"Migrating {old_template} to {new_quality_template}")
        shutil.copy2(old_template, new_quality_template)
        logger.info("Migration complete. You can now delete the old template.xlsx")

    # Check for quantity template
    quantity_template = config.template_dir / "zahvacene_kolicine_vode_template.xlsx"
    if not quantity_template.exists():
        # Look for it in various locations
        possible_locations = [
            Path("dnevni_ocevidnik_template.xlsx"),
            Path("zahvacene_kolicine_vode_template.xlsx"),
            Path("templates/dnevni_ocevidnik_template.xlsx")
        ]

        for location in possible_locations:
            if location.exists():
                logger.info(f"Found quantity template at {location}")
                shutil.copy2(location, quantity_template)
                break
        else:
            logger.warning(
                f"Quantity template not found! Please place "
                f"'zahvacene_kolicine_vode_template.xlsx' in {config.template_dir}"
            )


def main():
    """Main entry point."""
    try:
        # Load configuration
        config = Config.from_file()

        # Ensure directories exist
        config.export_dir.mkdir(parents=True, exist_ok=True)

        # Set up logging
        logger = setup_logging(config.export_dir)
        logger.info("Starting AquaExport Pro 2.1")

        # Create default config file if it doesn't exist
        if not Path("config.toml").exists():
            with open("config.toml", "w") as f:
                f.write('''[database]
host = "localhost"
port = 5432
name = "SCADA_arhiva_rab"
user = "postgres"
password = "fornax123"

[export]
directory = "./exports"
template_dir = "./templates"

# Water quality tag mappings
[tag_mappings.pk_barbat]
mutnoca = 3
klor = 21
temp = 134
pH = 132
redox = 133

[tag_mappings.vs_lopar]
klor = 151
temp = 155
redox = 156

[tag_mappings.vs_perici]
klor = 72
temp = 82
redox = 81
''')
            logger.info("Created default config.toml")

        # Migrate file structure if needed
        migrate_file_structure(config, logger)

        # Check templates exist
        quality_template = config.template_dir / "kvaliteta_vode_template.xlsx"
        quantity_template = config.template_dir / "zahvacene_kolicine_vode_template.xlsx"

        missing_templates = []
        if not quality_template.exists():
            missing_templates.append(quality_template.name)
        if not quantity_template.exists():
            missing_templates.append(quantity_template.name)

        if missing_templates:
            logger.error(f"Missing templates: {', '.join(missing_templates)}")
            messagebox.showerror(
                "Nedostaju predlošci",
                f"Excel predlošci nisu pronađeni!\n\n"
                f"Molimo postavite sljedeće datoteke u mapu {config.template_dir}:\n" +
                "\n".join(f"• {t}" for t in missing_templates)
            )
            return

        # Start GUI
        app = ModernDualModeGUI(config, logger)
        app.run()

    except Exception as e:
        if 'logger' in locals():
            logger.error(f"Application error: {e}\n{traceback.format_exc()}")
        messagebox.showerror(
            "Kritična greška",
            f"Aplikacija se ne može pokrenuti!\n\n{str(e)}"
        )


if __name__ == "__main__":
    main()